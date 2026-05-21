# セルに値を書き込む
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="富士 通太郎")
ws.cell(row=2, column=1, value="富士 明日花")
wb.save("名前一覧.xlsx")

