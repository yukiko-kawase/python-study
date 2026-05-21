import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="富士 太郎")
ws.cell(row=2, column=1, value="富士 明日花")

# ファイルパスを正しく指定できていない
wb.save(名前一覧.xlsx)