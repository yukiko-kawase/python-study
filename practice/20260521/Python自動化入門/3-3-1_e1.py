import openpyxl

wb = openpyxl.load_workbook("売上.xlsx")
wb.create_sheet(title=" 売上_4月")
# index=0を指定すると、シートが先頭に追加される
wb.create_sheet(title="利益_1月",index=0)

wb.save("売上(追加後).xlsx")