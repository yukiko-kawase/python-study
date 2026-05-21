"""
シートの追加
シートの追加には、Workbookオブジェクトが持つcreate_sheetメソッドを使用
キーワード引数のtitleには、追加するシート名を指定する
キーワード引数のindexでシートを追加する位置を指定できる
titleを指定しなかった場合、Sheet1、Sheet2のようにシート名が作成される
indexを指定しなかった場合、シートは最後に追加される
indexに0を指定すると、シートを先頭に追加される
"""
import openpyxl

wb = openpyxl.load_workbook("売上.xlsx")
wb.create_sheet(title=" 売上_4月")
wb.create_sheet(title="利益_1月",index=1)
wb.save("売上(追加後).xlsx")