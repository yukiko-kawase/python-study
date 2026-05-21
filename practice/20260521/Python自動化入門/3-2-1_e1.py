import openpyxl

wb = openpyxl.load_workbook("名簿.xlsx")
ws = wb.active
ws.insert_rows(1)
breakpoint()
ws.cell(row=1, column=1,value="姓")
ws.cell(row=1, column=2,value="名")
# 列の挿入する位置を文字列で指定しているとエラーになる
ws.insert_cols("B",amount=2)
breakpoint()
wb.save("名簿(挿入後).xlsx")