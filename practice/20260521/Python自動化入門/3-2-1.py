# 行や列の挿入
import openpyxl

# wb = openpyxl.Workbook()
# ws = wb.active
# sei_list = ["田中","山田","佐藤"]
# mei_list = ["太郎","花子","亮一"]

# row = 1
# for sei, mei in zip(sei_list, mei_list):
#     ws.cell(row=row, column=1, value=sei)
#     ws.cell(row=row, column=2, value=mei)
#     row += 1

# wb.save("名簿.xlsx")

wb = openpyxl.load_workbook("名簿.xlsx")
ws = wb.active
ws.insert_rows(1)
breakpoint()
ws.cell(row=1, column=1,value="姓")
ws.cell(row=1, column=2,value="名")
ws.insert_cols(2,amount=2)
breakpoint()
wb.save("名簿(挿入後).xlsx")






