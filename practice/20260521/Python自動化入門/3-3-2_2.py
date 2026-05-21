"""
シートをブックの末尾や戦闘に移動する
move_sheetメソッドでは、移動できる数よりも大きな数をキーワード引数offsetに指定した場合エラーにならない
そのため、ブックが持つシートの数だけ後ろに移動させれば、必ず末尾に移動させることができる
よって、sheetnamesプロパティで取得したリスト要素数をlen関数で求めると、シートの数が分かる
"""
import openpyxl

wb = openpyxl.load_workbook("売上.xlsx")
# offset=len(wb.sheetnames)と指定することでブックの末尾にシートを移動できる
wb.move_sheet("売上_1月", offset=len(wb.sheetnames))
#  offset=-len(wb.sheetnames)と指定することでブックの先頭にシートを移動できる
wb.move_sheet("売上_1月", offset=-len(wb.sheetnames))
wb.save("売上(移動後_2).xlsx")