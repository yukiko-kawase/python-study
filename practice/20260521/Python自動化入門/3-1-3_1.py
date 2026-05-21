import openpyxl
file_path = ("部署人数.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active
ws.cell(row=5,column=1, value="合計")
ws.cell(row=5, column=2, value="=SUM(B2:B4)")
wb.save("部署人数.xlsx")
