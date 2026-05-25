# 横棒グラフ作成
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference

wb = load_workbook("practice/20260525/売上.xlsx")
ws = wb[wb.sheetnames[0]]
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=4)
label = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=4)

bar = BarChart()
bar.type = "bar"
bar.add_data(data,titles_from_data=True)
bar.set_categories(label)
bar.x_axis.title = "品名"
bar.y_axis.title = "売上金額"
bar.title = "商品別売上金額"
ws.add_chart(bar,"D1")
wb.save("practice/20260525/売上(横棒グラフ作成後).xlsx")