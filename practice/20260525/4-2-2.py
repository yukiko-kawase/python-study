from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

wb = load_workbook("practice/20260525/経費.xlsx")
ws = wb.active
rule = CellIsRule(operator="greaterThan", formula= [10000], fill=PatternFill(patternType="solid",bgColor="FF0000"))
ws.conditional_formatting.add("D2:D4",rule)
wb.save("practice/20260525/経費(条件付き書式設定後).xlsx")