# 積み上げグラフの作成
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("practice/20260525/売上まとめ.xlsx")
ws = wb.active
data = Reference(ws,min_col=2,min_row=1, max_col=ws.max_column, max_row=ws.max_row)
label = Reference(ws,min_col=1,min_row=2, max_col=1, max_row=ws.max_row)
bar = BarChart()
bar.type = "col"
bar.grouping = "stacked"
bar.overlap = 100
bar.add_data(data,titles_from_data=True)
bar.set_categories(label)
bar.x_axis.title = "品名"
bar.y_axis.title = "売上金額"
ws.add_chart(bar,"F1")
wb.save("practice/20260525/売上まとめ(積み上げ棒グラフ作成後).xlsx")
