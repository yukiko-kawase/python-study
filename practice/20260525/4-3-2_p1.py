from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment,Border,PatternFill,Side

wb = load_workbook("practice/20260525/経費.xlsx")
ws = wb.active
ws.column_dimensions["A"].width = 16
rule = CellIsRule(operator="lessThanOrEqual",formula=[5000],fill=PatternFill(patternType="solid",fgColor="00FFFF"))
ws.conditional_formatting.add("D2:D4",rule)
pattern = PatternFill(patternType="solid",fgColor="B0B0B0")
for row in ws["A1:D1"]:
    for wc in row:
        wc.fill = pattern
        wc.alignment = Alignment(horizontal="center")
for row in ws["A2:A4"]:
    for wc in row:
        wc.number_format = "yyyy/mm/dd(aaa)"
for row in ws["D2:D4"]:
    for wc in row:
        wc.number_format = "#,###"
side = Side(style="thick")
Border_all = Border(top=side,bottom=side,left=side,right=side)
for row in ws["A1:D4"]:
    for wc in row:
        wc.border = Border_all