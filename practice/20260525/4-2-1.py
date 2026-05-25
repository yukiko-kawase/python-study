from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = load_workbook("practice/20260525/経費.xlsx")
ws = wb.active
ws.cell(row=1, column=5, value="状況")
validation = DataValidation(type="list", formula1='"未清算,清算中,清算済み"',showErrorMessage=True)
validation.add("E2:E4")
ws.add_data_validation(validation)
wb.save("practice/20260525/経費(入力規則設定後).xlsx")
