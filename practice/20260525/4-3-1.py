import unicodedata

import openpyxl

wb = openpyxl.load_workbook("practice/20260525/クラス名簿まとめ.xlsx")
Sheet_list = wb.worksheets
for ws in Sheet_list:
    ws.freeze_panes = "A2"
    ws["A1"].alignment = openpyxl.styles.Alignment(horizontal = "center")
    ws["B1"].alignment = openpyxl.styles.Alignment(horizontal = "center")
    for row in ws["A1:B4"]:
        for wc in row:
            wc.value = unicodedata.normalize("NFKC",wc.value)
wb.save("practice/20260525/クラス名簿まとめ(表記統一後).xlsx")
