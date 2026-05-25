# 円グラフ作成
from openpyxl import load_workbook
from openpyxl.chart import PieChart,Reference
# 各項目の割合を表示するクラスをインポート
from openpyxl.chart.label import DataLabelList


wb = load_workbook("practice/20260525/売上.xlsx")
ws = wb.active
data = Reference(ws,min_col=2, min_row=1, max_col=ws.max_column,max_row=ws.max_row)
label = Reference(ws,min_col=1, min_row=1, max_col=1,max_row=ws.max_row)
pie = PieChart()
pie.add_data(data,titles_from_data=True)
pie.set_categories(label)
pie.title = "商品別売上金額"
pie.dataLabels = DataLabelList(showPercent= True)
ws.add_chart(pie, "D1")
wb.save("practice/20260525/売上(円グラフ作成後).xlsx")

