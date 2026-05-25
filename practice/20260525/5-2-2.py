"""
グラフの書式設定
グラフのオブジェクトが持つlegend.position 属性に値を設定することで、凡例の位置を変更できる
凡例とはグラフのデータの色が人を表すのかを示す情報のこと
デフォルトではグラフ右に設定される
"""

# 棒グラフ作成
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.chart.series import DataPoint

wb = load_workbook("practice/20260525/売上.xlsx")
ws = wb[wb.sheetnames[0]]
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=4)
label = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=4)

bar = BarChart()
bar.type = "col"
bar.add_data(data,titles_from_data=True)
bar.set_categories(label)
bar.x_axis.title = "品名"
bar.y_axis.title = "売上金額"
bar.title = "商品別売上金額"
bar.legend.position = "l"
colors = ["B0B0B0", "00FFFF", "FF0000"]
for i in range(3):
    point = DataPoint(idx=i)
    point.graphicalProperties.solidFill = colors[i]
    # 棒グラフ一本目の凡例のデータポイントに色を代入
    bar.series[0].dPt.append(point)
    bar.y_axis.scaling.min = 2000
    bar.y_axis.scaling.max = 15000
ws.add_chart(bar,"D1")
wb.save("practice/20260525/売上(棒グラフ書式設定変更後).xlsx")
