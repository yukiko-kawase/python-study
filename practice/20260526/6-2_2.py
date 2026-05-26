import csv

from openpyxl import Workbook

wb = Workbook()
ws_a = wb.active
ws_a.title = "販売データ分析"
ws_b = wb.create_sheet(title="販売データ")
with open("販売関連データ\販売データ.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_b.append(row)
ws_c = wb.create_sheet(title="販売前在庫")
with open("販売関連データ\販売前在庫.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_c.append(row)
ws_d = wb.create_sheet(title="原価")
with open("販売関連データ\原価.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_d.append(row)
ws_b.column_dimensions["A"].width = 15
ws_c.column_dimensions["A"].width = 15
ws_d.column_dimensions["A"].width = 15

ws_a.column_dimensions["A"].width = 15
row_num = 1
for row in ws_b.values:
    col_num = 1
    for input_value in row:
        if row_num > 1 and col_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_value)
        col_num += 1
    row_num += 1
row_num = 1
col_num = 4
for row in ws_c["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1
row_num = 1
col_num = 5
for row in ws_d["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1
wb.save("販売データ分析.xlsx")