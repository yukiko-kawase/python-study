import csv

from openpyxl import Workbook

wb = Workbook()
ws_a = wb.active
ws_a.title = "販売データ分析"
ws_b = wb.create_sheet(title="販売データ")
with open("販売関連データ\販売データ.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_b.append(row)
ws_c = wb.create_sheet(title="販売前在庫")
with open("販売関連データ\販売前在庫.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_c.append(row)
ws_d = wb.create_sheet(title="原価")
with open("販売関連データ\原価.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_d.append(row)
ws_b.column_dimensions["A"].width = 15
ws_c.column_dimensions["A"].width = 15
ws_d.column_dimensions["A"].width = 15
wb.save("販売データ分析.xlsx")