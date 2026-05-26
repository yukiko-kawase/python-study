import csv

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, PatternFill, Side

wb = Workbook()
ws_a = wb.active
ws_a.title = "販売データ分析"
ws_b = wb.create_sheet(title="販売データ")
with open("販売関連データ\販売データ.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_b.append(row)
ws_c = wb.create_sheet(title="販売前在庫")
with open("販売関連データ\販売前在庫.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_c.append(row)
ws_d = wb.create_sheet(title="原価")
with open("販売関連データ\原価.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_d.append(row)
ws_b.column_dimensions["A"].width = 15
ws_c.column_dimensions["A"].width = 15
ws_d.column_dimensions["A"].width = 15

ws_a.column_dimensions["A"].width = 15
row_num = 1
for row in ws_b.values:
    col_num = 1
    for input_value in row:
        if row_num > 1 and col_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_value)
        col_num += 1
    row_num += 1
row_num = 1
col_num = 4
for row in ws_c["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1
row_num = 1
col_num = 5
for row in ws_d["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1

ws_a.insert_cols(5)
ws_a["E1"] = "在庫数(販売後)"
ws_a["G1"] = "売上金額"
ws_a["H1"] = "利益"
for row_num in range(2, 8):
    formula_1 = "=D" + str(row_num) + "-C" + str(row_num)
    formula_2 = "=B" + str(row_num) + "*C" + str(row_num)
    formula_3 = "=(B" + str(row_num) + "-F" + str(row_num) + ")*C" + str(row_num)
    ws_a.cell(row=row_num, column=5, value=formula_1)
    ws_a.cell(row=row_num, column=7, value=formula_2)
    ws_a.cell(row=row_num, column=8, value=formula_3)

ws_a.column_dimensions["E"].width = 15
ws_a.column_dimensions["G"].width = 10
ws_a.column_dimensions["H"].width = 10
pattern = PatternFill(patternType="solid", fgColor="B0B0B0")
side = Side(style="thin")
border_all = Border(top=side, bottom=side, left=side, right=side)
row_num = 1
for row in ws_a["A1:H7"]:
    col_num = 1
    for wc in row:
        wc.border = border_all
        if row_num == 1:
            wc.alignment = Alignment(horizontal="center")
            wc.fill = pattern
        elif row_num > 1 and col_num > 1:
            wc.number_format = "#,###"
        col_num += 1
    row_num += 1
rule = CellIsRule(operator="lessThan", formula=[100], fill=PatternFill(patternType="solid", bgColor="FF0000"))
ws_a.conditional_formatting.add("E2:E7", rule)

data_1 = Reference(ws_a, min_col=7, min_row=1, max_col=ws_a.max_column, max_row=ws_a.max_row)
label_1 = Reference(ws_a, min_col=1, min_row=2, max_col=1, max_row=ws_a.max_row)
bar_1 = BarChart()
bar_1.type = "col"
bar_1.add_data(data_1, titles_from_data=True)
bar_1.set_categories(label_1)
bar_1.x_axis.title = "商品名"
bar_1.y_axis.title = "売上金額"
bar_1.title = "商品別の売上金額と利益"
bar_1.width = 16.5
bar_1.height = 15
ws_a.add_chart(bar_1, "A9")
data_2 = Reference(ws_a, min_col=3, min_row=1, max_col=5, max_row=ws_a.max_row)
label_2 = Reference(ws_a, min_col=1, min_row=2, max_col=1, max_row=ws_a.max_row)
bar_2 = BarChart()
bar_2.type = "col"
bar_2.add_data(data_2, titles_from_data=True)
bar_2.set_categories(label_2)
bar_2.x_axis.title = "商品名"
bar_2.y_axis.title = "数量"
bar_2.title = "商品別の販売数と在庫数"
bar_2.width = 16.5
bar_2.height = 15
ws_a.add_chart(bar_2, "I9")
wb.save("販売データ分析.xlsx")