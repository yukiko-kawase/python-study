import csv
import datetime
import pathlib
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,PatternFill,Side,Font
from openpyxl.utils import get_column_letter

# csvファイルからExcelのブックへのデータ転記
wb_a = Workbook()
path = pathlib.Path("practice/20260526/売上データ_2月_第1週")
ws_a = wb_a.active
csv_list = sorted(path.glob("2月?日.csv"))
for name in csv_list:
    # basename()パスの最後のファイル名だけを取り出す
    ws_a.title = os.path.basename(name)[:-4]
    with open(name, "r", encoding="utf8") as file:
        row_num = 0
        for row in csv.reader(file):
            ws_a.append(row)
            row_num += 1
            if row_num ==1:
                ws_a.cell(row=row_num, column=5, value="売上金額(税抜)")
                ws_a.cell(row=row_num, column=6, value="売上金額(税込)")
            else:
                price = int(ws_a.cell(row=row_num, column=2).value)
                quantity = int(ws_a.cell(row=row_num, column=3).value)
                sales = price * quantity
                ws_a.cell(row=row_num, column=5, value=sales)
                if ws_a.cell(row=row_num, column=4,).value == "※":
                    ws_a.cell(row=row_num, column=6, value=sales*1.08)
                else:
                    ws_a.cell(row=row_num, column=6, value=sales*1.1)
    if len(wb_a.sheetnames) < len(csv_list):
        ws_a = wb_a.create_sheet()
# Excelの書式の設定
sheetlist = wb_a.worksheets
for ws in sheetlist:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
wb_a.save("practice/20260526/飲料売上(2月_第1週).xlsx")

# 売上レポートに表の作成とデータ転記
wb_b = Workbook()
ws_b = wb_b.active
ws_b.title = "売上金額(税抜)"
ws_b["A1"] = "商品名"
sheet_num = 1
for sheet in sheetlist:
    ws_b.cell(row=1, column=sheet_num+1, value=sheet.title)
    row_num = 1
    for row in sheet.values:
        col_num = 1
        for input_value in row:
            if row_num > 1 and col_num == 1 and sheet_num == 1:
                ws_b.cell(row=row_num, column=1, value=input_value)
            elif row_num > 1 and col_num == 5:
                ws_b.cell(row=row_num, column=sheet_num+1, value=input_value)
            col_num += 1
        row_num += 1
    sheet_num += 1

# 売上レポートに書式の設定
pattern = PatternFill(patternType="solid", fgColor="00FF00")
side = Side(style="thin")
border_all = Border(top=side, bottom=side, left=side, right=side)
range_all = get_column_letter(ws_b.min_column) + str(ws_b.min_row) + ":" + \
            get_column_letter(ws_b.max_column) + str(ws_b.max_row)
row_num = 1
for row in ws_b[range_all]:
    col_num = 1
    for wc in row:
        wc.border = border_all
        if row_num == 1:
            wc.alignment = Alignment(horizontal="center")
            wc.fill = pattern
            wc.font = Font(bold=True)
            if col_num == 1:
                ws_b.column_dimensions[get_column_letter(col_num)].width = 30
            else:
                ws_b.column_dimensions[get_column_letter(col_num)].width = 11
        elif row_num > 1 and col_num > 1:
            wc.number_format = "#,###"
        col_num += 1
    row_num += 1
ws_b.insert_rows(1, amount=5)
ws_b.cell(row=1, column=5, value="作成日")
ws_b["E1"].alignment = Alignment(horizontal="right")
ws_b.cell(row=1, column=6, value=datetime.datetime.now())
ws_b["F1"].number_format = "yyyy/m/d"
ws_b.cell(row=5, column=6, value="単位：円")
ws_b["F5"].alignment = Alignment(horizontal="right")
ws_b.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
ws_b.cell(row=3, column=1,value="飲料売上レポート(2月 第1週) ※税抜金額")
ws_b["A3"].font = Font(size=20)
ws["A3"].alignment = Alignment(horizontal="center")                
wb_b.save("practice/20260526/飲料売上レポート(2月_第1週).xlsx")

 
