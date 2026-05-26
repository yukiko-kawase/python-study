import csv
import pathlib
import os

from openpyxl import Workbook

# csvファイルからExcelのブックへのデータ転記
wb_a = Workbook()
path = pathlib.Path("practice/20260526/売上データ_2月_第1週")
ws_a = wb_a.active
csv_list = sorted(path.glob("2月?日.csv"))
for name in csv_list:
    # basename()パスの最後のファイル名だけを取り出す
    ws_a.title = os.path.basename(name)[:-4]
    with open(name, "r", encoding="utf8") as file:
        row_num = 0
        for row in csv.reader(file):
            ws_a.append(row)
            row_num += 1
            if row_num ==1:
                ws_a.cell(row=row_num, column=5, value="売上金額(税抜)")
                ws_a.cell(row=row_num, column=6, value="売上金額(税込)")
            else:
                price = int(ws_a.cell(row=row_num, column=2).value)
                quantity = int(ws_a.cell(row=row_num, column=3).value)
                sales = price * quantity
                ws_a.cell(row=row_num, column=5, value=sales)
                if ws_a.cell(row=row_num, column=4,).value == "※":
                    ws_a.cell(row=row_num, column=6, value=sales*1.08)
                else:
                    ws_a.cell(row=row_num, column=6, value=sales*1.1)
    if len(wb_a.sheetnames) < len(csv_list):
        ws_a = wb_a.create_sheet()
# Excelの書式の設定
sheetlist = wb_a.worksheets
for ws in sheetlist:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
wb_a.save("practice/20260526/飲料売上(2月_第1週).xlsx")

# 売上レポートに表の作成とデータ転記
wb_b = Workbook()
ws_b = wb_b.active
ws_b.title = "売上金額(税抜)"
ws_b["A1"] = "商品名"
sheet_num = 1
for sheet in sheetlist:
    ws_b.cell(row=1, column=sheet_num+1, value=sheet.title)
    row_num = 1
    for row in sheet.values:
        col_num = 1
        for input_value in row:
            if row_num > 1 and col_num == 1 and sheet_num == 1:
                ws_b.cell(row=row_num, column=1, value=input_value)
            elif row_num > 1 and col_num == 5:
                ws_b.cell(row=row_num, column=sheet_num+1, value=input_value)
            col_num += 1
        row_num += 1
    sheet_num += 1
wb_b.save("practice/20260526/飲料売上レポート(2月_第1週).xlsx")

 
