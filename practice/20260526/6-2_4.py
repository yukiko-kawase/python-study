import csv

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, PatternFill, Side

wb = Workbook()
ws_a = wb.active
ws_a.title = "販売データ分析"
ws_b = wb.create_sheet(title="販売データ")
with open("販売関連データ\販売データ.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_b.append(row)
ws_c = wb.create_sheet(title="販売前在庫")
with open("販売関連データ\販売前在庫.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_c.append(row)
ws_d = wb.create_sheet(title="原価")
with open("販売関連データ\原価.csv", "r", encoding="utf8") as file:
    for row in csv.reader(file):
        ws_d.append(row)
ws_b.column_dimensions["A"].width = 15
ws_c.column_dimensions["A"].width = 15
ws_d.column_dimensions["A"].width = 15

ws_a.column_dimensions["A"].width = 15
row_num = 1
for row in ws_b.values:
    col_num = 1
    for input_value in row:
        if row_num > 1 and col_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_value)
        col_num += 1
    row_num += 1
row_num = 1
col_num = 4
for row in ws_c["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1
row_num = 1
col_num = 5
for row in ws_d["B1:B7"]:
    for input_cell in row:
        if row_num > 1:
            ws_a.cell(row=row_num, column=col_num, value=int(input_cell.value))
        else:
            ws_a.cell(row=row_num, column=col_num, value=input_cell.value)
    row_num += 1

ws_a.insert_cols(5)
ws_a["E1"] = "在庫数(販売後)"
ws_a["G1"] = "売上金額"
ws_a["H1"] = "利益"
for row_num in range(2, 8):
    formula_1 = "=D" + str(row_num) + "-C" + str(row_num)
    formula_2 = "=B" + str(row_num) + "*C" + str(row_num)
    formula_3 = "=(B" + str(row_num) + "-F" + str(row_num) + ")*C" + str(row_num)
    ws_a.cell(row=row_num, column=5, value=formula_1)
    ws_a.cell(row=row_num, column=7, value=formula_2)
    ws_a.cell(row=row_num, column=8, value=formula_3)

ws_a.column_dimensions["E"].width = 15
ws_a.column_dimensions["G"].width = 10
ws_a.column_dimensions["H"].width = 10
pattern = PatternFill(patternType="solid", fgColor="B0B0B0")
side = Side(style="thin")
border_all = Border(top=side, bottom=side, left=side, right=side)
row_num = 1
for row in ws_a["A1:H7"]:
    col_num = 1
    for wc in row:
        wc.border = border_all
        if row_num == 1:
            wc.alignment = Alignment(horizontal="center")
            wc.fill = pattern
        elif row_num > 1 and col_num > 1:
            wc.number_format = "#,###"
        col_num += 1
    row_num += 1
rule = CellIsRule(operator="lessThan", formula=[100], fill=PatternFill(patternType="solid", bgColor="FF0000"))
ws_a.conditional_formatting.add("E2:E7", rule)
wb.save("販売データ分析.xlsx")