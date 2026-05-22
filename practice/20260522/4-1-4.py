from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("経費.xlsx")
ws = wb.active
pattern = PatternFill(patternType="solid", fgColor="00FFFF")
for row in ws["A1:D1"]:
    for wc in row:
        wc.fill = pattern
wb.save("経費(色設定後).xlsx")
