# ブックの保護
import pathlib
# load_workbook関数だけ呼び出したい時の書き方
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection

path = pathlib.Path(".")
for name in sorted(path.glob("売上_?月.xlsx")):
    # openxl記述不要
    wb = load_workbook(name)
    """
    ブックを保護するには、Workbookオブジェクトのsecurity属性にWorkboookProtectionクラスのオブジェクトが持つ
    workbookPassプロパティでにはパスワードを、lockstructure属性にはTrueを設定する
    """
    wb.security = WorkbookProtection(workbookPassword="fujitsu7825", lockStructure=True)
    wb.save("(保護済み)" + str(name))

