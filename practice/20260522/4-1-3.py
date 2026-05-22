from  openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook("経費.xlsx")
ws = wb.active
side = Side(style="thin")
border_all = Border(top=side, bottom=side, left=side, right=side)
for row in ws["A1:D4"]:
    for wc in row:
        wc.border = border_all
wb.save("経費(罫線設定後).xlsx")
