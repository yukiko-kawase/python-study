# セルの範囲を指定してデータを読み込む
import pathlib

import openpyxl

wb_a = openpyxl.load_workbook("売上_まとめ用.xlsx")
ws_a = wb_a.active
path = pathlib.Path(".")
col_num = 2
# ブックを切り替えて繰り返し処理
for name in sorted(path.glob("売上_?月.xlsx")):
    wb_b = openpyxl.load_workbook(name)
    ws_b = wb_b.active
    values = ws_b["B2:B4"]
    row_num = 2
    for row in values:
        for cell in row:
            input_value = cell.value
            ws_a.cell(row=row_num, column=col_num,value=input_value)
            row_num += 1
    col_num += 1
wb_a.save("売上_まとめ用(完了_2).xlsx")