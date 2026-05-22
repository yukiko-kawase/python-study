from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook("名簿.xlsx")
ws = wb.active
cells = ws["A1:B3"]
for row in cells:
    for ws_cell in row:
        ws_cell.alignment = Alignment(horizontal="center", vertical="top")
wb.save("名簿(配置変更).xlsx")