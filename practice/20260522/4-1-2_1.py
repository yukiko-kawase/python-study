from openpyxl import load_workbook

wb = load_workbook("経費.xlsx")
ws = wb.active
date_cells = ws["A2:A4"]
for row in date_cells:
    for wc in row:
        wc.number_format = "yyyy-mm-dd"
money_cells = ws["D2:D4"]
for row in money_cells:
    for wc in row:
        wc.number_format = "#,###"
wb.save("経費(表示形式変更後).xlsx")