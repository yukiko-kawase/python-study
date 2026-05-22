import openpyxl

wb = openpyxl.load_workbook("英字の文字列.xlsx")
ws = wb.active
ws["C2"].value = ws["C2"].value.upper()
ws["C3"].value = ws["C3"].value.lower()
ws["C4"].value = ws["C4"].value.title()
ws["C5"].value = ws["C5"].value.title()
ws["C6"].value = ws["C6"].value.swapcase()
wb.save("英字の文字列(変換後).xlsx")

