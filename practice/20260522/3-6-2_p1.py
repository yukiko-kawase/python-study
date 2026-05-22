import csv

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "売上_1月"
with open("売上_1月.csv", "r", encoding="utf8") as file:
    row_num = 1
    for row in csv.reader(file):
        column_num = 1
        for value in row:
            if row_num > 1 and column_num == 2:
                ws.cell(row=row_num, column=column_num, value=int(value))
            else:
                ws.cell(row=row_num, column=column_num, value=value)
            column_num += 1
        row_num += 1
for month in range(2,13):
    wb.create_sheet(title="売上_" + str(month) +"月")
wb.save("売上CSVまとめ.xlsx")

