from openpyxl import load_workbook

wb = load_workbook("日付の表示形式.xlsx")
ws = wb.active

ws["C2"].number_format  = "yyyy/m/d"
ws["C3"].number_format  = "yyyy/mm/dd"
ws["C4"].number_format  = "yyyy/m/d(ddd)"
ws["C5"].number_format  = "yyyy/m/dddd"
ws["C6"].number_format  = 'yyyy"年"m"月"d"日"'
ws["C7"].number_format  = 'yyyy"年"mm"月"dd"日"'
ws["C8"].number_format  = 'ggge"年"m"月"d"日"'
ws["C9"].number_format  = 'ggge"年"m"月"d"日"'
ws["C10"].number_format = 'ggge"年"m"月"d"日"(aaa)'
ws["C11"].number_format = 'm"月"d"日"(aaa)'
ws["C12"].number_format = 'm"月"d"日" aaaa'

wb.save("日付の表示形式(設定後).xlsx")
