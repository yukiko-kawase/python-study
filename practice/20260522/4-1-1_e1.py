from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook("名簿.xlsx")
ws = wb.active
cells = ws["A1:B3"]
# Worksheetオブジェクトのvaluesプロパティで取得した値は、セルに入力されている値(文字列)だけを取得する
# そのためセルに入力されている値の配置は、設定できない
# 6行目のcellsはセル自体のデータを格納しているので、値の配置を設定できる
for row in ws.values:
    for ws_cell in row:
        ws_cell.alignment = Alignment(horizontal="center", vertical="top")
wb.save("名簿(配置変更).xlsx")