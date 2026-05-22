from openpyxl import load_workbook

wb = load_workbook("数値の表示形式.xlsx")
ws = wb.active
ws["C2"].number_format = "#,##0"
ws["C3"].number_format = "#,##0"
ws["C4"].number_format = "#,###"
ws["C5"].number_format = "#,###"
ws["C6"].number_format = "0.000"
ws["C7"].number_format = "0.000"
ws["C8"].number_format = "#.###"
ws["C9"].number_format = "#.###"
wb.save("数値の表示形式(設定後).xlsx")
