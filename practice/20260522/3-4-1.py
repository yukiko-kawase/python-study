import openpyxl

wb_a = openpyxl.load_workbook("売上.xlsx")
ws_a = wb_a.active
wb_b = openpyxl.load_workbook("利益.xlsx")
ws_b = wb_b.active
input_value = ws_a.cell(row=2, column=2).value
ws_b.cell(row=2, column=2, value=input_value)
ws_b.cell(row=2, column=4, value="=B2-C2")
wb_b.save("利益(転記後).xlsx")


