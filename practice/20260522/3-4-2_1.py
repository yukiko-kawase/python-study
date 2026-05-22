import pathlib

import openpyxl

wb_a = openpyxl.load_workbook("売上_まとめ用.xlsx")
ws_a = wb_a.active
path = pathlib.Path(".")
col_num = 2
for name in sorted(path.glob("売上_?月.xlsx")):
    wb_b = openpyxl.load_workbook(name)
    ws_b = wb_b.active
    input_value = ws_b.cell(row=2, column=2).value
    ws_a.cell(row=2, column=col_num, value=input_value)
    col_num += 1
    wb_a.save("売上_まとめ用(完了).xlsx")