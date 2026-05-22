#文字設定の変更とセルの結合
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

wb = load_workbook("経費.xlsx")
ws = wb.active
ws.insert_rows(1,amount=2)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws.cell(row=1, column=1, value="経費実績一覧")
ws["A1"].font = Font(size=15)
ws["A1"].alignment = Alignment(horizontal="center")
wb.save("経費(文字変更後とセル結合).xlsx")
